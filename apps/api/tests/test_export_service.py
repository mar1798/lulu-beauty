import io

import openpyxl

from app.export.service import (
    HEADER,
    TOTAL_LABEL,
    OrderExportRow,
    build_orders_workbook,
    content_disposition,
    readable_filename_label,
    sanitize_filename_label,
)


def _row(**overrides: object) -> OrderExportRow:
    defaults: dict[str, object] = {
        "product_name": "Rose Serum",
        "brand": "Lulu",
        "quantity": 2,
        "unit_price_cents": 1500,
    }
    defaults.update(overrides)
    return OrderExportRow(**defaults)  # type: ignore[arg-type]


def _sheet(rows: list[OrderExportRow], *, include_prices: bool = True):  # type: ignore[no-untyped-def]
    content = build_orders_workbook(rows, include_prices=include_prices)
    workbook = openpyxl.load_workbook(io.BytesIO(content))
    sheet = workbook.active
    assert sheet is not None
    return sheet


def _values(sheet, index: int) -> list[object]:  # type: ignore[no-untyped-def]
    return [cell.value for cell in next(sheet.iter_rows(min_row=index, max_row=index))]


def test_build_orders_workbook_writes_header_row() -> None:
    sheet = _sheet([])

    assert sheet.title == "Заказ"
    assert _values(sheet, 1) == HEADER


def test_build_orders_workbook_writes_a_line_per_product_with_the_line_total() -> None:
    sheet = _sheet([_row(product_name="Крем", brand="Lulu", quantity=3, unit_price_cents=15000)])

    assert _values(sheet, 2) == ["Крем", "Lulu", 3, 150.0, 450.0]


def test_build_orders_workbook_closes_with_the_totals_row() -> None:
    rows = [
        _row(product_name="Крем", quantity=3, unit_price_cents=15000),
        _row(product_name="Тоник", quantity=2, unit_price_cents=5000),
    ]

    sheet = _sheet(rows)

    assert sheet.max_row == 4  # шапка + два товара + итог
    # Пустые ячейки под брендом и ценой openpyxl читает как None.
    assert _values(sheet, 4) == [TOTAL_LABEL, None, 5, None, 550]


def test_build_orders_workbook_totals_row_on_an_empty_export_is_zeroed() -> None:
    sheet = _sheet([])

    assert _values(sheet, 2) == [TOTAL_LABEL, None, 0, None, 0]


def test_build_orders_workbook_without_prices_drops_both_money_columns() -> None:
    rows = [_row(product_name="Крем", brand="Lulu", quantity=3, unit_price_cents=15000)]

    sheet = _sheet(rows, include_prices=False)

    assert _values(sheet, 1) == HEADER[:3]
    assert _values(sheet, 2) == ["Крем", "Lulu", 3]
    assert _values(sheet, 3) == [TOTAL_LABEL, None, 3]


def test_build_orders_workbook_without_prices_keeps_the_quantity_format() -> None:
    """Regression guard: the formats list is filtered alongside the columns, not by index."""
    sheet = _sheet([_row(quantity=1200)], include_prices=False)

    quantity_cell = next(sheet.iter_rows(min_row=2, max_row=2))[2]
    assert quantity_cell.number_format == "#,##0"


def test_build_orders_workbook_widens_columns_to_the_longest_value() -> None:
    """Regression: fixed widths clipped long product names into the neighbouring column."""
    short = _sheet([_row(product_name="Крем")])
    long = _sheet([_row(product_name="Крем для лица с гиалуроновой кислотой, 50 мл")])

    assert long.column_dimensions["A"].width > short.column_dimensions["A"].width
    # Шапка длиннее любого бренда в примере — колонка всё равно не уже неё.
    assert short.column_dimensions["B"].width > len(HEADER[1])


def test_build_orders_workbook_caps_the_text_columns_and_wraps_instead() -> None:
    sheet = _sheet([_row(product_name="Крем " * 60)])

    assert sheet.column_dimensions["A"].width < 100
    name_cell = next(sheet.iter_rows(min_row=2, max_row=2))[0]
    assert name_cell.alignment.wrap_text is True


def test_sanitize_filename_label_keeps_alphanumeric_dash_underscore() -> None:
    assert sanitize_filename_label("June 2026 / Batch #1") == "June-2026---Batch--1"


def test_sanitize_filename_label_passthrough_for_safe_label() -> None:
    assert sanitize_filename_label("new-year-batch_2") == "new-year-batch_2"


def test_sanitize_filename_label_drops_non_ascii() -> None:
    """Regression: Cyrillic is alnum in Python, so it used to survive into the header."""
    sanitized = sanitize_filename_label("Июнь 2030")
    assert sanitized.isascii()
    assert sanitized.endswith("2030")
    assert sanitize_filename_label("june-2030") == "june-2030"


def test_readable_label_keeps_cyrillic_but_drops_path_characters() -> None:
    assert readable_filename_label("Июнь 2030") == "Июнь-2030"
    assert readable_filename_label("a/b:c") == "a-b-c"


def test_content_disposition_is_latin1_encodable_and_keeps_extension() -> None:
    header = content_disposition("orders-Июнь-2030.xlsx")

    # Starlette encodes response headers as latin-1; a non-encodable value was a 500.
    header.encode("latin-1")
    assert 'filename="orders-' in header
    assert '2030.xlsx"' in header  # the extension survives the ASCII fallback
    assert "filename*=UTF-8''" in header
    assert "%D0%98%D1%8E%D0%BD%D1%8C" in header  # percent-encoded "Июнь"


def test_content_disposition_handles_ascii_filename_unchanged() -> None:
    header = content_disposition("orders-all-cycles.xlsx")
    assert 'filename="orders-all-cycles.xlsx"' in header
