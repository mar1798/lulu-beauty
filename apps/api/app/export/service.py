import functools
import io
import uuid
from dataclasses import dataclass
from urllib.parse import quote

import anyio.to_thread
from openpyxl import Workbook
from openpyxl.cell import Cell, WriteOnlyCell
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sqlalchemy import func, literal, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.catalog.models import Product
from app.config import settings
from app.cycles.models import OrderCycle
from app.orders.models import Order, OrderItem, OrderStatus

HEADER = [
    "Наименование товара",
    "Бренд",
    "Количество",
    f"Цена за штуку, {settings.currency}",
    f"Сумма, {settings.currency}",
]

# Money columns are optional: the same sheet is used both as the owner's own purchase list
# and as something handed to a supplier, and the latter has no business seeing what the shop
# charges. Dropping them is a column filter over the full row, so nothing else has to know.
MONEY_COLUMNS = (3, 4)

TOTAL_LABEL = "Итого"

MONEY_FORMAT = "#,##0.00"
QUANTITY_FORMAT = "#,##0"

# Excel's column width unit is about one character of the default font, so the widths below
# are derived from the longest rendered value in each column. Text columns stop growing at
# MAX_TEXT_WIDTH and wrap instead — an unbounded width would push the money columns off the
# screen on a single 255-character product name.
WIDTH_PADDING = 3
MAX_TEXT_WIDTH = 60
TEXT_COLUMNS = (0, 1)

CellValue = str | float | None


@dataclass(frozen=True)
class OrderExportRow:
    """One product line of the purchase list: the same product across every order, summed."""

    product_name: str
    brand: str
    quantity: int
    unit_price_cents: int

    @property
    def total_cents(self) -> int:
        return self.unit_price_cents * self.quantity


def _money(cents: int) -> float:
    return cents / 100


def _rendered_width(value: CellValue, number_format: str | None) -> int:
    """Length of the text Excel will actually paint in the cell."""
    if isinstance(value, float) and number_format == MONEY_FORMAT:
        return len(f"{value:,.2f}")
    if isinstance(value, int) and number_format == QUANTITY_FORMAT:
        return len(f"{value:,}")
    return len(str(value))


def _visible[T](values: list[T], *, include_prices: bool) -> list[T]:
    """Drops the unit-price and line-total cells when the export is asked for without money.

    Generic on purpose: the same filter runs over the header, the number formats and every
    body row, so the money columns are described in exactly one place (MONEY_COLUMNS).
    """
    if include_prices:
        return values
    return [value for index, value in enumerate(values) if index not in MONEY_COLUMNS]


def build_orders_workbook(rows: list[OrderExportRow], *, include_prices: bool = True) -> bytes:
    """Serializes the purchase list to xlsx bytes. Blocking CPU — call it off the event loop.

    write_only, because the default Workbook keeps every cell as a Python object until
    save() and then walks the lot: on a full-catalogue export (100k lines) that was ~7.6s
    of the request, against ~2.3s of actual SQL. In write-only mode openpyxl streams each
    row to the underlying archive as it is appended, which is both far faster and flat in
    memory. The trade-off is that the sheet cannot be read back or re-ordered afterwards,
    which no caller here does — hence the column widths being computed up front, before the
    first append, rather than measured while writing.
    """
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet(title="Заказ")

    all_formats: list[str | None] = [None, None, QUANTITY_FORMAT, MONEY_FORMAT, MONEY_FORMAT]
    formats = _visible(all_formats, include_prices=include_prices)
    header = _visible(HEADER, include_prices=include_prices)
    total_quantity = sum(row.quantity for row in rows)
    total_cents = sum(row.total_cents for row in rows)

    body: list[list[CellValue]] = [
        _visible(
            [
                row.product_name,
                row.brand,
                row.quantity,
                _money(row.unit_price_cents),
                _money(row.total_cents),
            ],
            include_prices=include_prices,
        )
        for row in rows
    ]
    footer: list[CellValue] = _visible(
        [TOTAL_LABEL, "", total_quantity, "", _money(total_cents)],
        include_prices=include_prices,
    )

    widths = [len(str(title)) for title in header]
    for values in (*body, footer):
        for index, value in enumerate(values):
            widths[index] = max(widths[index], _rendered_width(value, formats[index]))

    for index, width in enumerate(widths):
        limit = MAX_TEXT_WIDTH if index in TEXT_COLUMNS else width
        letter = get_column_letter(index + 1)
        sheet.column_dimensions[letter].width = min(width, limit) + WIDTH_PADDING

    def cell(value: CellValue, *, index: int, bold: bool = False) -> Cell:
        written = WriteOnlyCell(sheet, value=value)
        # A product name is text, and openpyxl infers otherwise: it writes any string
        # starting with "=" as a real <f> formula. Product names come from the catalog
        # import, and this sheet is explicitly meant to be forwarded to a supplier — so a
        # row named "=1+1" (or something that reads a cell, or worse) would arrive as a
        # live formula in somebody else's Excel. Declaring the type keeps the text exactly
        # as typed, with none of the leading apostrophe the usual workaround leaves behind.
        # Only "=" is affected: "+", "-" and "@" are already written as strings, and in
        # xlsx — unlike csv — Excel does not re-interpret those.
        if isinstance(value, str):
            written.data_type = "s"
        written.alignment = Alignment(
            wrap_text=index in TEXT_COLUMNS,
            vertical="center",
            horizontal="left" if index in TEXT_COLUMNS else "right",
        )
        number_format = formats[index]
        if number_format is not None:
            written.number_format = number_format
        if bold:
            written.font = Font(bold=True)
        return written

    sheet.append(
        [cell(title, index=index, bold=True) for index, title in enumerate(header)],
    )
    for values in body:
        sheet.append([cell(value, index=index) for index, value in enumerate(values)])
    sheet.append([cell(value, index=index, bold=True) for index, value in enumerate(footer)])

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def sanitize_filename_label(label: str) -> str:
    """ASCII-only, filename-safe rendering of a label.

    str.isalnum() is True for Cyrillic, so an earlier version let non-ASCII through into
    Content-Disposition, where Starlette encodes headers as latin-1 — any Russian cycle
    label turned the export into a 500. Non-ASCII is dropped here and the readable name is
    carried separately by the RFC 5987 filename* parameter (see content_disposition).
    """
    return "".join(
        char if (char.isascii() and char.isalnum()) or char in "-_" else "-" for char in label
    )


def readable_filename_label(label: str) -> str:
    """Keeps letters of any alphabet, drops only what is unsafe inside a filename."""
    unsafe = set('/\\:*?"<>|') | {chr(code) for code in range(32)}
    return "".join("-" if char in unsafe or char == " " else char for char in label)


def content_disposition(filename: str) -> str:
    """attachment header with an ASCII fallback plus the UTF-8 name (RFC 5987).

    Old clients read filename=; anything current prefers filename*, so the owner still
    downloads "orders-Июнь-2030.xlsx" rather than a row of dashes.
    """
    stem, _, extension = filename.rpartition(".")
    if stem:
        ascii_name = f"{sanitize_filename_label(stem)}.{extension}"
    else:
        ascii_name = sanitize_filename_label(filename)
    return f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{quote(filename, safe='')}"


class ExportService:
    def __init__(self, session: AsyncSession) -> None:
        self._session = session

    async def export_orders(
        self,
        cycle_id: uuid.UUID | None,
        status: OrderStatus | None = None,
        *,
        include_prices: bool = True,
    ) -> tuple[bytes, str]:
        rows = await self._export_rows(cycle_id, status, include_prices=include_prices)

        # In a worker thread: everything above is I/O the event loop can interleave, but
        # the xlsx serialization is pure CPU, and on a full export it ran long enough
        # (seconds) to stall every other request in the process — the API is single-
        # threaded, so a blocking call here is downtime for the whole shop, not just for
        # the owner waiting on their download.
        content = await anyio.to_thread.run_sync(
            functools.partial(build_orders_workbook, rows, include_prices=include_prices),
        )
        filename = await self._filename(cycle_id)
        return content, filename

    async def _export_rows(
        self,
        cycle_id: uuid.UUID | None,
        status: OrderStatus | None = None,
        *,
        include_prices: bool = True,
    ) -> list[OrderExportRow]:
        """The purchase list: every ordered product summed across the orders in scope.

        Grouped in SQL rather than in Python, so the sheet is one row per product no matter
        how many customers asked for it. The grouping key includes the snapshotted price:
        order items keep the price they were bought at, and two orders placed either side of
        a price change are genuinely two lines on the supplier's list, not one line with an
        invented average.

        That last part only holds while the price is on the sheet. With include_prices off
        the price drops out of the grouping key too — otherwise the same product bought at
        two prices would come out as two visually identical lines, which reads as a bug to
        whoever gets the file. The price then stays 0 on the row: it is not written anywhere,
        and carrying one of the two real prices would make the row quietly wrong.

        Brand is not snapshotted on the item, so it comes from a LEFT JOIN on products —
        soft-deleted products keep their row, but an item whose product_id went NULL (a hard
        delete) still belongs on the list, with the same "—" the admin listing shows.
        """
        quantity = func.sum(OrderItem.quantity)
        price = OrderItem.product_price_cents if include_prices else literal(0)
        query = (
            select(
                OrderItem.product_name,
                Product.brand,
                price.label("unit_price_cents"),
                quantity,
            )
            .select_from(OrderItem)
            .join(Order, Order.id == OrderItem.order_id)
            .outerjoin(Product, Product.id == OrderItem.product_id)
            .group_by(OrderItem.product_name, Product.brand)
            .order_by(Product.brand.nulls_last(), OrderItem.product_name)
        )
        if include_prices:
            query = query.group_by(OrderItem.product_price_cents)
        if cycle_id is not None:
            query = query.where(Order.cycle_id == cycle_id)
        if status is not None:
            query = query.where(Order.status == status)

        result = await self._session.execute(query)
        return [
            OrderExportRow(
                product_name=product_name,
                brand=brand if brand else "—",
                quantity=int(total_quantity),
                unit_price_cents=unit_price_cents,
            )
            for (product_name, brand, unit_price_cents, total_quantity) in result.all()
        ]

    async def _filename(self, cycle_id: uuid.UUID | None) -> str:
        if cycle_id is None:
            return "orders-all-cycles.xlsx"

        cycle = await self._session.get(OrderCycle, cycle_id)
        label = cycle.label if cycle is not None and cycle.label else str(cycle_id)
        return f"orders-{readable_filename_label(label)}.xlsx"
