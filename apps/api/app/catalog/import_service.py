import csv
import io
import re
import uuid
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from pathlib import Path

import anyio.to_thread
import openpyxl
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.catalog.models import Category, Product
from app.catalog.schemas import SLUG_PATTERN, ImportRowErrorResponse, ImportSummaryResponse
from app.catalog.service import UNIQUE_VIOLATION
from app.common.limits import MAX_PRICE_CENTS, MAX_VOLUME_ML
from app.orders.service import OrderPriceChange, OrdersService

REQUIRED_HEADERS = {"name", "slug", "price"}
# `normalize_header` only lowercases and turns spaces/dashes into underscores, so the
# camelCase spelling the admin panel documents ("inStock") arrives as "instock" and used
# to miss `row["in_stock"]` entirely — every row silently imported as in stock.
# `volume_ml`/`объем` are what a supplier's price list actually calls the column the
# admin panel documents as "volume" — all three mean the same millilitres.
HEADER_ALIASES = {
    "instock": "in_stock",
    "volume_ml": "volume",
    "объем": "volume",
    "объём": "volume",
}
# The columns are String(255) in the database; past that the row failed at flush, which is
# a 500 for the whole upload rather than one reported line.
MAX_TEXT_LENGTH = 255
# Ceiling on a cell's decimal exponent, checked before the value is turned into an int.
# Well above any real price or volume (both are bounded far lower a few lines later) and
# far below the point where `int()` on a Decimal becomes a denial of service.
_MAX_DECIMAL_EXPONENT = 18
# A file may be well under MAX_IMPORT_BYTES and still be enormous: xlsx is deflate, and a
# 4 MB upload of 300 000 rows expands to ~150 MB of parsed rows alone, before the ORM
# objects. Rows are refused as a file error, like an unreadable file — there is no useful
# per-row report to give for "this file is too big".
MAX_IMPORT_ROWS = 50_000
# How many bad rows the report carries back. A file that is wrong in every line would
# otherwise put a message per line into one JSON response.
MAX_REPORTED_ERRORS = 200
# How many slugs are looked up per round-trip when resolving a file against the catalogue.
_SLUG_CHUNK = 5000
TRUTHY_VALUES = {"true", "1", "yes", "y", "да"}
FALSY_VALUES = {"false", "0", "no", "n", "нет"}

# Practical transliteration, not GOST: a slug has to be readable, not reversible.
# Mirrors packages/widgets/src/utils/slug.ts, which builds slugs for the same columns
# from the admin forms.
_TRANSLITERATION = {
    "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "e", "ж": "zh",
    "з": "z", "и": "i", "й": "y", "к": "k", "л": "l", "м": "m", "н": "n", "о": "o",
    "п": "p", "р": "r", "с": "s", "т": "t", "у": "u", "ф": "f", "х": "h", "ц": "ts",
    "ч": "ch", "ш": "sh", "щ": "sch", "ъ": "", "ы": "y", "ь": "", "э": "e", "ю": "yu",
    "я": "ya",
}  # fmt: skip


class ImportFileError(Exception):
    pass


class ImportRowError(Exception):
    pass


def normalize_header(raw: str | None) -> str:
    normalized = (raw or "").strip().lower().replace(" ", "_").replace("-", "_")
    return HEADER_ALIASES.get(normalized, normalized)


def slugify(value: str) -> str:
    """A `SLUG_PATTERN`-shaped slug from arbitrary text (may be empty if nothing survives)."""
    transliterated = "".join(_TRANSLITERATION.get(char, char) for char in value.lower())
    return re.sub(r"[^a-z0-9]+", "-", transliterated).strip("-")[:255].strip("-")


def category_name_from_slug(slug: str) -> str:
    """`"eye-cream"` → `"Eye Cream"` — a placeholder name for a category the file invented.

    The file only carries a slug, and a category has to be named something the owner can
    recognise in the admin list; it is renamed there in one field if this reads badly.
    """
    return " ".join(word.capitalize() for word in slug.split("-"))


def parse_price_cents(raw: str | None) -> int:
    text = (raw or "").strip()
    if not text:
        raise ImportRowError("не указана цена")
    try:
        value = Decimal(text)
    except InvalidOperation as error:
        raise ImportRowError(f"цена не распознана: {raw!r}") from error
    # Decimal() happily parses "nan" and "inf". Both then blow up further down —
    # NaN raises on the comparison, Infinity on int() — and an exception here is not a
    # row error: it escapes import_file() and turns the whole upload into a 500, losing
    # every valid row in the file along with the report of what was wrong.
    if not value.is_finite():
        raise ImportRowError(f"цена не распознана: {raw!r}")
    if value < 0:
        raise ImportRowError("цена не может быть отрицательной")
    # Checked on the exponent, *before* any arithmetic: `Decimal("1E+1000000")` is finite
    # and non-negative, so it passes both guards above — and then `int()` on it spends
    # forty seconds materialising a million digits on the event loop before raising
    # `decimal.Overflow`, which is not an ImportRowError and takes the whole upload down
    # as a 500. `adjusted()` is the decimal exponent and costs nothing to read.
    if value.adjusted() > _MAX_DECIMAL_EXPONENT:
        raise ImportRowError(f"слишком большая цена: {raw!r}")

    cents = int((value * 100).to_integral_value(rounding=ROUND_HALF_UP))
    # price_cents is a 32-bit column; past that the row fails at flush, i.e. again as a
    # 500 for the whole file rather than as one reported line.
    if cents > MAX_PRICE_CENTS:
        raise ImportRowError(f"слишком большая цена: {raw!r}")
    return cents


def parse_volume_ml(raw: str | None) -> int | None:
    """`"50"`, `"50 мл"`, `"50,5"` → millilitres, or None for an empty cell.

    Rounded rather than refused: a price list writes 50.5 ml as often as 50, and the
    volume is a caption on a card, not money. Zero and negatives *are* refused — «0 мл»
    on a card says something false, and the empty cell already means "unknown".
    """
    text = (raw or "").strip().lower().removesuffix("мл").removesuffix("ml").strip()
    if not text:
        return None
    try:
        value = Decimal(text.replace(",", "."))
    except InvalidOperation as error:
        raise ImportRowError(f"объём не распознан: {raw!r}") from error
    if not value.is_finite():
        raise ImportRowError(f"объём не распознан: {raw!r}")
    # Same trap as in `parse_price_cents`, same reason — see the note there.
    if value.adjusted() > _MAX_DECIMAL_EXPONENT:
        raise ImportRowError(f"слишком большой объём: {raw!r}")

    millilitres = int(value.to_integral_value(rounding=ROUND_HALF_UP))
    if millilitres <= 0:
        raise ImportRowError(f"объём должен быть больше нуля: {raw!r}")
    if millilitres > MAX_VOLUME_ML:
        raise ImportRowError(f"слишком большой объём: {raw!r}")
    return millilitres


def parse_in_stock(raw: str | None) -> bool:
    text = (raw or "").strip().lower()
    if not text:
        return True
    if text in TRUTHY_VALUES:
        return True
    if text in FALSY_VALUES:
        return False
    raise ImportRowError(f"непонятное значение в колонке «в наличии»: {raw!r}")


def parse_csv_rows(content: bytes) -> list[tuple[int, dict[str, str]]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as error:
        raise ImportFileError("файл не в кодировке UTF-8") from error

    reader = csv.DictReader(io.StringIO(text))
    rows: list[tuple[int, dict[str, str]]] = []
    # `csv.Error` is raised lazily, from the iteration rather than the constructor, and the
    # likeliest cause is mundane: one unclosed quote in a price list makes the reader
    # swallow the rest of the file as a single field and trip the 128k field limit. That
    # is a bad file, not a 500.
    try:
        if reader.fieldnames is None:
            return []
        for row_number, row in enumerate(reader, start=2):
            if len(rows) >= MAX_IMPORT_ROWS:
                raise _too_many_rows()
            rows.append((row_number, {normalize_header(k): (v or "") for k, v in row.items()}))
    except csv.Error as error:
        raise ImportFileError(f"не удалось прочитать csv-файл: {error}") from error
    return rows


def parse_xlsx_rows(content: bytes) -> list[tuple[int, dict[str, str]]]:
    # The whole body is inside the guard, not just `load_workbook`: `read_only=True` means
    # the sheet's XML is parsed lazily, during `iter_rows`, so a corrupt sheet in an
    # otherwise valid archive opened fine and then raised `ParseError` from the loop —
    # past the point where it could still become the reported "не удалось прочитать".
    try:
        return _read_xlsx_rows(content)
    except ImportFileError:
        raise
    except Exception as error:  # noqa: BLE001 - openpyxl raises various errors for a corrupt file
        raise ImportFileError(f"не удалось прочитать xlsx-файл: {error}") from error


def _read_xlsx_rows(content: bytes) -> list[tuple[int, dict[str, str]]]:
    workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    sheet = workbook.active
    if sheet is None:
        return []

    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []
    headers = [normalize_header(str(h) if h is not None else "") for h in header_row]

    rows: list[tuple[int, dict[str, str]]] = []
    for row_number, values in enumerate(rows_iter, start=2):
        # Seeded with every header, so a short row still reports which columns the *file*
        # has. Which columns are present is what decides whether an absent value means
        # "leave it alone" or "clear it" (see `_validate_row`), and csv.DictReader answers
        # the same question the same way — a missing cell is a key with an empty value.
        row = dict.fromkeys(headers, "")
        row.update(
            {
                headers[i]: ("" if v is None else str(v))
                for i, v in enumerate(values)
                if i < len(headers)
            }
        )
        if any(value for value in row.values()):
            if len(rows) >= MAX_IMPORT_ROWS:
                raise _too_many_rows()
            rows.append((row_number, row))
    return rows


def _too_many_rows() -> ImportFileError:
    return ImportFileError(
        f"в файле больше {MAX_IMPORT_ROWS} строк — разделите его на несколько файлов"
    )


def parse_rows(filename: str, content: bytes) -> list[tuple[int, dict[str, str]]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        return parse_csv_rows(content)
    if suffix in (".xlsx", ".xlsm"):
        return parse_xlsx_rows(content)
    raise ImportFileError(f"неподдерживаемый тип файла: {suffix or 'без расширения'}")


def validate_headers(rows: list[tuple[int, dict[str, str]]]) -> str | None:
    if not rows:
        return None
    missing = REQUIRED_HEADERS - set(rows[0][1].keys())
    if missing:
        return f"в файле нет обязательных колонок: {', '.join(sorted(missing))}"
    return None


class CategoryIndex:
    """The catalogue's categories, plus the ones the file invents as it is read.

    A file is written by a supplier or by hand, not against the category list, so a
    `category` cell naming something the shop does not have yet used to fail the whole
    row. It now creates the category instead: the column stays optional, and an empty
    cell still means "no category".
    """

    def __init__(self, session: AsyncSession, categories: list[Category]) -> None:
        self._session = session
        self._by_slug = {category.slug: category for category in categories}
        self._by_name = {category.name.strip().lower(): category for category in categories}
        # New categories land after everything the owner has already ordered by hand.
        highest = max((category.sort_order for category in categories), default=-1)
        self._next_sort_order = highest + 1

    def resolve(self, raw: str) -> Category | None:
        """The category for a `category` cell, creating it when the catalogue has none."""
        text = raw.strip()
        if not text:
            return None

        # The cell is a slug when it looks like one, and a name otherwise — files come
        # both ways, and a Russian name is never a valid slug.
        slug = text if re.fullmatch(SLUG_PATTERN, text) else slugify(text)
        existing = self._by_slug.get(slug) or self._by_name.get(text.lower())
        if existing is not None:
            return existing
        if not slug:
            raise ImportRowError(f"категория не распознана: {text!r}")

        # `slugify` truncates, the pass-through branch does not — and `categories.slug` is
        # String(255), so a 300-character cell that happens to *look* like a slug reached
        # the flush and came back as StringDataRightTruncationError: a 500 with no machine
        # code, the whole file rolled back, and no line number to point at.
        if len(slug) > MAX_TEXT_LENGTH:
            raise ImportRowError(f"категория длиннее {MAX_TEXT_LENGTH} символов")

        name = text if slug != text else category_name_from_slug(slug)
        category = Category(name=name[:255], slug=slug, sort_order=self._next_sort_order)
        self._next_sort_order += 1
        self._session.add(category)
        self._by_slug[slug] = category
        self._by_name[name.lower()] = category
        return category


class CatalogImportService:
    def __init__(self, session: AsyncSession) -> None:
        self._session = session

    async def import_file(
        self, filename: str, content: bytes
    ) -> tuple[ImportSummaryResponse, list[OrderPriceChange]]:
        """Applies a price list, and reports the orders its prices moved.

        The price changes come back rather than being announced here: the customers are
        told over Telegram, which must not happen until the transaction the caller owns
        has actually committed.
        """
        try:
            # Off the event loop: openpyxl parsing a multi-megabyte upload is seconds of
            # blocking CPU, and the API is single-threaded — see export/service.py.
            rows = await anyio.to_thread.run_sync(parse_rows, filename, content)
        except ImportFileError as error:
            return ImportSummaryResponse(
                created=0, updated=0, errors=[ImportRowErrorResponse(row=0, message=str(error))]
            ), []

        header_error = validate_headers(rows)
        if header_error is not None:
            return ImportSummaryResponse(
                created=0, updated=0, errors=[ImportRowErrorResponse(row=0, message=header_error)]
            ), []

        categories = CategoryIndex(self._session, await self._load_categories())
        brands_by_key = await self._load_brands()
        # Which columns the file actually carries. An absent column must leave the existing
        # value alone: a supplier price list is name/slug/price and nothing else, and
        # writing the defaults for everything it omits wiped the brand, description and
        # category off every product it touched — and put them all back in stock.
        columns = set(rows[0][1]) if rows else set()

        created = 0
        updated = 0

        # Off the event loop, like the parsing above. The loop is pure CPU over every row
        # of the file, and this process serves the shop, the bot and the scheduler from
        # one thread — a big price list held all three for as long as validation took.
        validated, errors = await anyio.to_thread.run_sync(
            self._validate_rows, rows, columns, categories, brands_by_key
        )

        # Every product the file could touch, in one query rather than one per row. The
        # per-row SELECT also forced an autoflush each time, so a 2 000-row upload cost
        # ~4 000 round-trips and grew linearly (2.2s, and a full catalogue far worse).
        existing = await self._load_products_by_slug(
            [str(fields["slug"]) for fields in validated]
        )

        # Taken before `_upsert` overwrites them: after the loop the ORM objects carry the
        # file's prices, and "did this row change anything" is no longer answerable.
        previous_prices = {slug: product.price_cents for slug, product in existing.items()}

        for fields in validated:
            if self._upsert(existing, fields):
                created += 1
            else:
                updated += 1

        try:
            await self._session.flush()
        except IntegrityError as error:
            # A slug this file creates can be created by hand in another tab a moment
            # earlier — `_slug_conflict_as_error` describes exactly this race for the
            # single-product path. Reported as a line in the summary rather than as a 500.
            if getattr(error.orig, "sqlstate", None) != UNIQUE_VIOLATION:
                raise
            # The whole file goes: the flush is one statement batch, and everything it
            # queued is still pending in the session — the router's commit would simply
            # replay the same conflict. A savepoint would not help for the same reason.
            await self._session.rollback()
            return ImportSummaryResponse(
                created=0,
                updated=0,
                errors=[
                    ImportRowErrorResponse(
                        row=0,
                        message=(
                            "во время импорта кто-то создал товар или категорию с таким же "
                            "slug — повторите импорт"
                        ),
                    )
                ],
            ), []

        # A price list is not only a catalog edit. Orders still awaiting confirmation
        # quote these products, and until this ran, the mass path silently left them on
        # the old price while the hand-edit path (`PATCH /admin/products/{id}`) pulled it
        # through — the owner paid one number and saw another. Only prices that actually
        # moved, and only for products that existed before this file.
        moved: dict[uuid.UUID, int] = {}
        for slug in dict.fromkeys(str(fields["slug"]) for fields in validated):
            product = existing.get(slug)
            was = previous_prices.get(slug)
            if product is None or was is None or was == product.price_cents:
                continue
            moved[product.id] = product.price_cents

        # One call for the whole file, not one per product: a price list moves thousands
        # of prices, and a round trip each would hold this process — which also runs the
        # bot and the scheduler — for the length of the import.
        changes = await OrdersService(self._session).reprice_products(moved)

        return ImportSummaryResponse(created=created, updated=updated, errors=errors), changes

    def _validate_rows(
        self,
        rows: list[tuple[int, dict[str, str]]],
        columns: set[str],
        categories: CategoryIndex,
        brands_by_key: dict[str, str],
    ) -> tuple[list[dict[str, object]], list[ImportRowErrorResponse]]:
        """Every row's fields, and the report for the ones that have none.

        The error list is capped: a file whose every line is wrong (a mis-mapped column,
        a price list in another currency) would otherwise put one message per line into a
        single JSON response. The tail is summarised rather than dropped silently.
        """
        validated: list[dict[str, object]] = []
        errors: list[ImportRowErrorResponse] = []
        suppressed = 0

        for row_number, row in rows:
            try:
                validated.append(self._validate_row(row, columns, categories, brands_by_key))
            except ImportRowError as error:
                if len(errors) < MAX_REPORTED_ERRORS:
                    errors.append(ImportRowErrorResponse(row=row_number, message=str(error)))
                else:
                    suppressed += 1

        if suppressed:
            errors.append(
                ImportRowErrorResponse(row=0, message=f"…и ещё {suppressed} строк с ошибками")
            )
        return validated, errors

    def _validate_row(
        self,
        row: dict[str, str],
        columns: set[str],
        categories: CategoryIndex,
        brands_by_key: dict[str, str],
    ) -> dict[str, object]:
        """The fields one line writes — only the ones its file actually has a column for.

        `columns` is what keeps a partial file partial: `_upsert` assigns whatever comes
        back from here, so a key present with a default is indistinguishable from a value
        the owner meant to set.
        """
        name = row.get("name", "").strip()
        if not name:
            raise ImportRowError("не указано название")
        if len(name) > MAX_TEXT_LENGTH:
            raise ImportRowError(f"название длиннее {MAX_TEXT_LENGTH} символов")

        slug = row.get("slug", "").strip()
        if not slug:
            raise ImportRowError("не указан slug")
        if len(slug) > MAX_TEXT_LENGTH:
            raise ImportRowError(f"slug длиннее {MAX_TEXT_LENGTH} символов")
        if not re.fullmatch(SLUG_PATTERN, slug):
            raise ImportRowError(f"недопустимый slug: {slug!r} — только латиница, цифры и дефис")

        fields: dict[str, object] = {
            "name": name,
            "slug": slug,
            "price_cents": parse_price_cents(row.get("price")),
        }

        if "volume" in columns:
            fields["volume_ml"] = parse_volume_ml(row.get("volume"))
        if "in_stock" in columns:
            fields["in_stock"] = parse_in_stock(row.get("in_stock"))
        if "description" in columns:
            fields["description"] = row.get("description", "").strip() or None

        if "brand" in columns:
            # Optional, unlike in the product form: a file is what the shop is stocked
            # from, and one nameless brand is not a reason to reject the row it sits on.
            # Such a product is invisible to the catalog's brand filter until the brand
            # is filled in.
            brand: str | None = row.get("brand", "").strip() or None
            if brand is not None:
                if len(brand) > MAX_TEXT_LENGTH:
                    raise ImportRowError(f"бренд длиннее {MAX_TEXT_LENGTH} символов")
                # One spelling per brand, whatever case the file happens to use: the brand
                # column has no table behind it, so "round lab" would otherwise become a
                # second brand alongside "Round Lab" and split the catalog filter in two.
                # New brands register themselves, so case variants inside a single upload
                # collapse onto the first row that named it.
                brand = brands_by_key.setdefault(brand.lower(), brand)
            fields["brand"] = brand

        if "category" in columns:
            # Last, so a row rejected further up never leaves a category behind with no
            # products in it. The relationship rather than `category_id`: a category this
            # file has just invented has no id until flush.
            fields["category"] = categories.resolve(row.get("category", ""))

        return fields

    def _upsert(self, existing: dict[str, Product], fields: dict[str, object]) -> bool:
        """Upserts by slug; returns True if a new product was created, False if updated.

        `existing` is both the lookup table and the record of what this file has already
        created: a slug repeated inside one upload has to update the row the earlier line
        produced, not insert a second one against a UNIQUE column.
        """
        slug = str(fields["slug"])
        product = existing.get(slug)

        if product is None:
            product = Product(**fields)
            self._session.add(product)
            existing[slug] = product
            return True

        for field, value in fields.items():
            setattr(product, field, value)
        product.deleted_at = None  # re-importing revives a previously discontinued product
        return False

    async def _load_products_by_slug(self, slugs: list[str]) -> dict[str, Product]:
        """The products these rows will update, keyed by slug — chunked, not one big IN.

        A bind parameter per slug against a file that may carry tens of thousands of rows
        would run into the driver's parameter ceiling; the chunk size is well under it.
        """
        products: dict[str, Product] = {}
        unique = list(dict.fromkeys(slugs))
        for start in range(0, len(unique), _SLUG_CHUNK):
            result = await self._session.execute(
                select(Product).where(Product.slug.in_(unique[start : start + _SLUG_CHUNK]))
            )
            products.update({product.slug: product for product in result.scalars().all()})
        return products

    async def _load_categories(self) -> list[Category]:
        result = await self._session.execute(select(Category))
        return list(result.scalars().all())

    async def _load_brands(self) -> dict[str, str]:
        """Brands the catalogue already knows, keyed by lowercase spelling.

        Soft-deleted products count: a brand left only on discontinued rows is
        still the spelling the catalogue uses, and re-importing it under another
        case would revive it as a second brand.
        """
        result = await self._session.execute(
            select(Product.brand).where(Product.brand.is_not(None), Product.brand != "").distinct()
        )
        return {brand.lower(): brand for brand in result.scalars().all() if brand is not None}
